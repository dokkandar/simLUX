#!/usr/bin/env python
"""Build the AutoRASM Settings Guide (Excel) with a sketch next to each
sketchable variable. The variable list + ORDER come straight from
cad_app/src/varreg.rs (same source as the settings menu) so they always match;
plain-English explanations + sketches are joined on top.

SAMPLE scope below = whichever sections are listed in SECTIONS.
Run:  python tools/build_guide.py
"""
import os, re, math
from PIL import Image, ImageDraw
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PNG  = os.path.join(ROOT, "tools", "sketch_png")
REG  = os.path.join(ROOT, "cad_app", "src", "varreg.rs")
os.makedirs(PNG, exist_ok=True)

SECTIONS = ["Display & Visual Feedback", "Selection & Grips", "Object Snaps & Precision", "Editing & Behavior", "File & Save", "Xrefs & Images", "UI & Workspace", "Plot & Publish", "System & Performance", "View & Navigation", "Miscellaneous", "RUST_CAD-specific", "Code-Audit Hardcoded", "Grid & CARD", "UCS Icon", "Drafting Defaults"]   # expand as we walk each section

# ---- parse varreg.rs (authoritative order) --------------------------------
def parse_varreg():
    out = []
    for line in open(REG, encoding="utf-8"):
        if not re.match(r'\s*Var \{ name:', line):
            continue
        def g(p, d=""):
            m = re.search(p, line); return m.group(1) if m else d
        name = g(r'name:\s*"([^"]*)"')
        sec  = g(r'section:\s*"([^"]*)"')
        desc = g(r'desc:\s*"([^"]*)"')
        stat = g(r'status:\s*Status::(\w+)')
        dflt = g(r'default:\s*"([^"]*)"')
        wired = g(r'wired:\s*(true|false)') == "true"
        kraw = g(r'kind:\s*(.*),\s*status:')
        out.append(dict(name=name, sec=sec, desc=desc, status=stat,
                        default=dflt, wired=wired, kraw=kraw))
    return out

def kind_type_options(kraw):
    if kraw.startswith("Kind::Bool"):  return "Bool", "on/off", None
    if kraw.startswith("Kind::Color"): return "Colour", "colour", None
    if kraw.startswith("Kind::Text"):  return "Text", "text", None
    m = re.match(r'Kind::U8\s*\{\s*min:\s*(\d+),\s*max:\s*(\d+)', kraw)
    if m: return "Number", f"{m.group(1)}-{m.group(2)}", None
    m = re.match(r'Kind::Int\s*\{\s*min:\s*([\d_]+),\s*max:\s*([\d_]+)', kraw)
    if m: return "Number", f"{int(m.group(1))}-{int(m.group(2)):,}", None
    if kraw.startswith("Kind::Float"): return "Number", "number", None
    m = re.match(r'Kind::Choice\(&\[(.*)\]\)', kraw)
    if m:
        opts = [o.strip().strip('"') for o in m.group(1).split(",")]
        return "Choice", " / ".join(opts), opts
    return "?", "", None

def default_display(dflt, opts):
    if opts is not None:
        try: return opts[int(dflt)]
        except Exception: return dflt
    if dflt == "true":  return "on"
    if dflt == "false": return "off"
    return dflt

# ---- plain-English explanations (the value-add, keyed by name) ------------
# what / options-override / notes.  Missing -> falls back to varreg desc.
EXPLAIN = {
 "CrsHrS": ("Crosshair cursor length, as % of screen (small cross to full-window lines).", "1-100 %", ""),
 "DrDspM": ("Show a moving ghost of an object while you drag it in Move/Copy.", None, ""),
 "HltSel": ("Selected objects light up in a highlight colour.", None, ""),
 "SelPrv": ("Pre-highlight an object when you hover over it, before clicking.", None, ""),
 "RllTp":  ("Show a tooltip with an object's info when you hover over it.", None, ""),
 "SelAr":  ("Selection box shows a translucent colour fill (not just an outline).", None, ""),
 "SelPrvL":("Max number of objects to pre-highlight at once (so big selections don't lag).", None, ""),
 "WinACol":("Fill colour of the WINDOW box (blue, dragged left to right).", None, ""),
 "CrsACol":("Fill colour of the CROSSING box (green, dragged right to left).", None, ""),
 "IntsCol":("Colour of the intersection markers where objects cross.", None, ""),
 "IntsDsp":("Show the intersection markers where objects cross.", None, ""),
 "LyLkFd": ("How much to fade objects on a LOCKED layer (0 = none, 100 = almost invisible).", "0-100 %", ""),
 "LodAnc": ("In simplified draft display, which point of an object the dot sits on.", None, ""),
 "WpFrmM": ("Wipeout (mask) border: off / on / only when selected.", None, ""),
 "AperBx": ("Show the snap-target box around the crosshair (size = SpTGSZ).", None, ""),
 "TrnDsp": ("Draw objects' transparency (see-through) level on/off.", None, ""),
 "RvClCrM":("How revision clouds (the change-marking lumpy outline) are drawn.", "mode", ""),
 "LnFade": ("Dims (fades) lines while you are editing.", None, ""),
 "PrvFlt": ("Object types skipped during the hover preview-highlight.", None, ""),
 "TrkPth": ("Show the dotted tracking / alignment guide lines.", None, ""),
 "BkgPlt": ("Print/plot in the background while you keep working.", None, "no printing here"),
 "GalVw":  ("Thumbnail gallery for blocks.", None, "no blocks gallery"),
 "HpQckP": ("Live quick-preview of a hatch fill before committing.", None, ""),
 "ImgHlt": ("Highlight a raster image's frame on hover/select.", None, "no images"),
 "LtGlyD": ("Show light-source icons (lighting design).", None, "no lights"),
 "MTxtFx": ("Fixed-width editor box for multi-line text.", None, ""),
 "OleHid": ("Hide embedded OLE objects (e.g. a pasted-in live Excel table).", None, "none here"),
 "PcBnd":  ("Show a 3D point cloud's bounding box.", None, "2D app"),
 "PcClpF": ("Show the crop-boundary outline of a 3D point cloud (laser scan).", None, "2D app"),
 "RvClGrp":("Grips on revision clouds.", None, ""),
 "TryIco": ("Show the app's icon in the Windows system tray (by the clock).", None, ""),
 "TryTim": ("How long tray notifications stay on screen (seconds).", None, ""),
 "WmfBkg": ("Background colour for WMF (old Windows clip-art) export.", None, ""),
 "WmfFrg": ("Foreground colour for WMF export.", None, ""),

 # --- Selection & Grips ---
 "GrClrS": ("Colour of a 'hot' grip - the square you've clicked and are about to drag.", "colour", ""),
 "GrClrU": ("Colour of normal (cool) grips - the little squares shown on a selected object.", "colour", ""),
 "GrpBlk": ("Show grips on each part inside a block, not just one grip at its insertion point.", None, ""),
 "GrpEnb": ("Master on/off for grips - the little editing squares that appear on a selected object.", None, ""),
 "GrpObjL":("If you select more than this many objects, grips are hidden (keeps editing fast).", None, ""),
 "GrpSz":  ("Size of the grip squares, in pixels.", "1-20 px", ""),
 "GrpTip": ("Show a small tooltip when you hover over a grip.", None, ""),
 "HidTxt": ("Hide text while you move or rotate it, so it doesn't clutter the drag preview.", None, ""),
 "ObjIsoM":("Isolate the selected objects - temporarily hide everything else to work cleanly.", None, ""),
 "OsnNdLg":("Use the old (legacy) behaviour for the Node object-snap.", None, "snap setting - belongs in Object Snaps"),
 "OsnOpt": ("Which object-snap points are active (a bitmask of end/mid/centre/...).", "bitmask", "snap setting - belongs in Object Snaps"),
 "PkAdd":  ("Each new pick ADDS to the selection. Off = each pick replaces it (Shift to add).", None, ""),
 "PkAuto": ("Click empty space and drag to start a selection window automatically.", None, ""),
 "PkDrag": ("How you draw a selection window: press-drag-release, or click then click.", None, ""),
 "PkFrst": ("Noun/verb editing: pick the objects FIRST, then run a command on them.", None, ""),
 "SelCyc": ("When objects overlap, cycle through them to pick the one underneath.", None, ""),
 "SelOfSc":("Allow selecting objects that are currently scrolled off-screen.", None, ""),
 "SubSelM":("Select a sub-part (one edge / vertex / face) instead of the whole object.", None, ""),
 "SelDmTm":("How long you must hold the mouse before a drag becomes a selection-window drag.", "50-2000 ms", ""),
 "GrpHvR": ("How close (pixels) the cursor must get to a grip before it can be grabbed.", "4-80 px", ""),

 # --- Object Snaps & Precision ---
 "OsnCrd": ("Let a typed coordinate override a running object-snap (keyboard wins over the snap).", None, ""),
 "PkBxSz": ("Size of the little pick box at the cursor used to click-select objects, in pixels.", "1-40 px", ""),
 "PolAdA": ("Extra polar-tracking angles on top of the main one (e.g. 30, 45), as a list.", "angle list", ""),
 "PolAng": ("Base polar-tracking angle the cursor locks to (90 = horizontal/vertical, 45 = diagonals too).", "0-360 deg", ""),
 "PolDst": ("Step distance when polar SNAP is on - the cursor jumps along the ray in these increments.", "units", ""),
 "PolMod": ("Turn polar tracking on - the cursor snaps to set angles as you draw.", None, ""),
 "SpTGSZ": ("Size of the object-snap target box (the aperture) shown around snap points, in pixels.", "4-80 px", ""),
 "TmpOvr": ("Allow temporary override keys - hold a key to force or suppress a snap for just one pick.", None, ""),

 # --- Editing & Behavior ---
 "AtDlgM": ("When inserting a block with attributes, pop a dialog to fill them in (vs typing each one).", None, ""),
 "AtPrmM": ("Prompt for a block's attribute values while inserting it.", None, ""),
 "BActBM": ("Show the action bar in the Block Editor.", None, "block editor - n/a here"),
 "BlkEdLk":("Lock the Block Editor so its contents can't be changed.", None, "block editor - n/a"),
 "BlkEdtr":("Internal flag: whether the Block Editor is currently open.", None, "block editor state"),
 "BlkMrL": ("How many recently-used blocks to keep in the quick list.", None, "blocks"),
 "BndTyp": ("How an external reference is bound into the drawing (bind vs insert).", None, "xref feature"),
 "CmDlgM": ("Use dialog boxes for commands like PLOT, instead of command-line prompts.", None, ""),
 "DblClkE":("Double-click an object to open its edit action.", None, ""),
 "EdgMod": ("Let Trim/Extend use an object's imaginary extension as the cutting/boundary edge.", None, ""),
 "HpMaxA": ("Largest hatch area that still gets a live preview (above this, preview is skipped).", None, "hatch"),
 "HpObjW": ("Warn before making a hatch that touches more than this many objects.", None, "hatch"),
 "HpSep":  ("Create a separate hatch object for each closed area, instead of one combined hatch.", None, "hatch"),
 "InpHMd": ("Show recently-typed values in the dynamic-input box as you draw.", None, ""),
 "MTjigS": ("The sample text shown in the live preview while you place multi-line text.", "text", ""),
 "PedAcc": ("Skip the 'convert to polyline?' question in PEDIT - just assume yes.", None, ""),
 "PrsPul": ("How the Presspull tool behaves.", None, "3D feature"),
 "RefPtTp":("How paths to referenced files are stored: full / relative / none.", None, "xref paths"),
 "SavFid": ("Save annotative objects so they display correctly in older programs.", None, "annotative/legacy"),
 "SbyLyr": ("Force new objects' colour/linetype/etc. to 'ByLayer' automatically.", None, ""),
 "SrfAsc": ("Keep surfaces linked to the curves they were built from.", None, "3D surfaces"),
 "TblInd": ("Show the row/column header indicators while editing a table.", None, "tables"),
 "TblTbr": ("Show the floating toolbar while editing a table.", None, "tables"),
 "XEdit":  ("Allow this drawing to be edited in place when another drawing references it.", None, "xref feature"),
 "XFdCtl": ("Fade the surrounding objects while you edit a reference in place.", None, "xref feature"),

 # --- File & Save ---
 "AudCtl": ("Write an audit report to a file when checking a drawing for errors.", None, ""),
 "AutoPub":("Automatically publish/export the drawing whenever you save or close it.", None, ""),
 "DgnMpP": ("File that maps layers/styles when importing DGN (MicroStation) drawings.", "path", "DGN import"),
 "DwgChk": ("Warn when opening a DWG that wasn't created by genuine Autodesk software.", None, "DWG compat"),
 "IsvBak": ("Keep a .bak backup copy each time you save.", None, ""),
 "IsvPrc": ("When wasted space in the file passes this %, do a full clean save instead of incremental.", "0-100 %", ""),
 "LogFlM": ("Write a log file recording your session.", None, ""),
 "LogFlP": ("Folder where the session log file is written.", "path", ""),
 "OpnPrt": ("Allow opening just part of a very large drawing (partial open).", None, ""),
 "RcovMd": ("Try to automatically recover the drawing after a crash.", None, ""),
 "SavFP":  ("Folder where automatic-save files are written.", "path", ""),
 "SavTim": ("How often (in minutes) to auto-save your work.", "minutes", ""),
 "SigWarn":("Warn about digital signatures when opening a file.", None, "signatures"),
 "SldChk": ("Validate 3D solids when a drawing loads.", None, "3D feature"),
 "TrstPth":("Folders treated as trusted for loading code and support files.", "path", ""),

 # --- Xrefs & Images ---
 "XrLdMd": ("How referenced drawings load: fully, on-demand, or on-demand via a local copy.", None, "xref feature"),
 "XrTmpP": ("Folder for the temporary local copies of referenced (xref) files.", "path", "xref feature"),
 "XrCtl":  ("Write a log file of xref attach/reload activity.", None, "xref feature"),
 "XrLyr":  ("Default layer that an attached xref's own layers nest under.", "layer", "xref feature"),
 "XrNtfy": ("Notify you when a referenced drawing changes on disk.", None, "xref feature"),
 "XrTyp":  ("Default attachment type for new xrefs: attach or overlay.", None, "xref feature"),
 "XdwFd":  ("Fade a referenced drawing so it stands out from your own work (0 = none).", "0-100 %", "xref feature"),
 "RastDpi":("Resolution (DPI) used when plotting raster images.", None, "raster images"),
 "RastPrc":("Share of memory allowed for holding raster images.", "0-100 %", "raster images"),
 "RastThr":("Memory threshold before raster images get paged out.", None, "raster images"),
 "OleQlty":("Print quality for embedded OLE objects.", None, "OLE feature"),
 "OleStrt":("Launch the source app when loading an OLE object.", None, "OLE feature"),
 "PdfShx": ("Treat SHX text in an imported PDF as real, editable text.", None, "PDF import"),
 "PdfShxL":("Layer to place SHX text recognized from a PDF.", "layer", "PDF import"),

 # --- UI & Workspace ---
 "DobMenu":("Custom menu / UI-definition file to load on startup.", "path", "legacy CUI"),
 "LokUI":  ("Lock toolbars and palettes so they can't be moved or resized.", None, ""),
 "MnuBar": ("Show the classic text menu bar (File / Edit / View ...) across the top.", None, ""),
 "MnuCtl": ("Show the old-style screen menu down the side.", None, "legacy"),
 "NavBar": ("Show the navigation bar (pan / zoom / orbit shortcuts).", None, "n/a here"),
 "NavCube":("Show the 3D ViewCube in the corner.", None, "3D feature"),
 "PalOpq": ("How see-through docked palettes are (100 = solid).", "0-100 %", ""),
 "QpLoc":  ("Where the Quick Properties panel appears.", None, "quick-props"),
 "QpMod":  ("Turn the Quick Properties pop-up on/off.", None, "quick-props"),
 "RibSta": ("Whether the ribbon starts minimized.", None, "ribbon"),
 "ScrnBx": ("Show the legacy screen-menu boxes.", None, "legacy"),
 "ShctMn": ("Show right-click shortcut (context) menus.", None, ""),
 "StartUp":("Show a startup dialog when the app or a new drawing opens.", None, ""),
 "TbCust": ("Allow customizing toolbars by dragging buttons around.", None, ""),
 "TltEnb": ("Show a tooltip when you hover over a toolbar / ribbon button.", None, ""),
 "TltMrg": ("Merge the basic and extended tooltips into one popup.", None, ""),
 "TltTrn": ("How see-through tooltips are.", "0-100 %", ""),
 "TpPalP": ("Folder where tool palettes are stored.", "path", ""),
 "TxtEd":  ("External text-editor application for editing text or scripts.", "app/path", ""),

 # --- Plot & Publish ---
 "PapUpd": ("Warn when a layout's paper size needs updating for the chosen printer.", None, "plotting - n/a"),
 "PStPlc": ("Default plot-style mode for new drawings (colour-dependent vs named).", None, "plotting - n/a"),
 "PubAll": ("When publishing, include all sheets by default.", None, "publishing - n/a"),
 "PubHch": ("Include hatch fills when publishing.", None, "publishing - n/a"),

 # --- System & Performance ---
 "FlDlgM": ("Use simple command-line file prompts instead of the file-browser dialog.", None, ""),
 "FntAlt": ("Substitute font to use when a drawing's font isn't installed.", "font", "fonts"),
 "FntMap": ("File that maps missing fonts to installed replacements.", "path", "fonts"),
 "LspAsD": ("Load acad.lsp automatically into every drawing.", None, "LISP - n/a"),
 "MxActVp":("Maximum number of viewports drawn at once.", None, "viewports"),
 "MxSort": ("Largest list that still gets sorted alphabetically (bigger lists stay unsorted, for speed).", None, ""),
 "PrxNot": ("Notify when a drawing contains proxy objects from a missing app.", None, "proxy objects"),
 "PrxShw": ("How proxy objects are displayed (show / hide / bounding box).", None, "proxy objects"),
 "PrxWeb": ("Offer a web link to get the missing app for proxy objects.", None, "proxy objects"),
 "StdViol":("Notify when something breaks the drawing standards.", None, "CAD standards"),
 "SysMon": ("Watch key system variables and warn when one changes.", None, ""),
 "TreMax": ("Memory limit (KB) for the spatial-index tree that speeds up redraws.", None, ""),
 "TxtFil": ("Fill text solid. Off = show outlines only (faster on big drawings).", None, ""),
 "TxtQlt": ("Smoothness of TrueType text rendering (higher = smoother, slower).", None, ""),
 "UntMod": ("How units / coordinates are displayed (decimal, architectural, ...).", None, ""),
 "WhipArc":("On-screen smoothness of arcs and circles (higher = smoother, slower).", None, ""),

 # --- View & Navigation ---
 "GeoLoc": ("Show the geographic-location marker (real-world coordinates).", None, "geo - n/a"),
 "LayTab": ("Show the Model / Layout tabs along the bottom.", None, "layouts"),
 "RtDsp":  ("Redraw continuously while panning/zooming (smooth) instead of only at the end.", None, ""),
 "StepSz": ("Step distance for Walk / Fly navigation.", "units", "3D navigation"),
 "StpPrSc":("Steps per second in Walk / Fly navigation.", None, "3D navigation"),
 "SunPrW": ("Show the Sun Properties window.", None, "3D lighting"),
 "UcsOrt": ("Auto-switch the UCS to match an orthographic view.", None, "UCS"),
 "VtDur":  ("How long a smooth view transition lasts (milliseconds).", "ms", ""),
 "VtEnbl": ("Animate view changes smoothly instead of jumping instantly.", None, ""),
 "VtFps":  ("Target frame rate for smooth view transitions.", "FPS", ""),
 "VwUpdA": ("Automatically refresh the view when the drawing changes.", None, ""),
 "ZmFact": ("How much each mouse-wheel notch zooms.", "factor", ""),
 "ZmWhl":  ("Direction of mouse-wheel zoom (forward = zoom in, or out).", None, ""),

 # --- Miscellaneous ---
 "Chrma":  ("How colour books (named colour palettes) are displayed.", None, "colour books"),
 "LyrDlgM":("Whether the Layer Properties manager is a dialog or a dockable panel.", None, "layers"),
 "LyrFlA": ("Warn when a layer filter would hide a large number of layers.", None, "layers"),
 "LyrNtf": ("Notify when new layers are added (e.g. by an xref).", None, "layers"),
 "MTxtEd": ("External editor application for multi-line text.", "app/path", ""),
 "PrjNam": ("Search path used to locate project / referenced files.", "path", ""),
 "SsmAuto":("Auto-open the Sheet Set Manager when a sheet set is loaded.", None, "sheet sets - n/a"),
 "SsmPol": ("How often (seconds) the Sheet Set Manager checks for changes.", None, "sheet sets - n/a"),
 "SsmSta": ("Internal flag: Sheet Set Manager open/closed.", None, "sheet sets - n/a"),

 # --- RUST_CAD-specific ---
 "GpuRnd": ("How drawing is rendered: CPU, automatic GPU, or forced GPU.", None, ""),
 "FpsDsp": ("Show a frames-per-second overlay (performance check).", None, ""),
 "IdxDsp": ("Show the spatial-index status overlay (debug / performance).", None, ""),
 "IdxCel": ("Tuning: target number of index cells per object (affects pick & redraw speed).", "number", ""),
 "BgCol":  ("Canvas background colour.", "colour", ""),
 "SnpPri": ("Order snaps are preferred when several are nearby (e.g. end before mid).", "snap list", ""),
 "SnpAct": ("Which object snaps are switched on at startup (a bitmask).", "bitmask", ""),
 "TabCyc": ("Press Tab to cycle through nearby snap candidates.", None, ""),
 "CmdEcho":("Echo typed commands into the history / log.", None, ""),
 "CmdHisM":("How many command-history lines to keep.", "lines", ""),
 "RubBnd": ("Style of the rubber-band line while drawing: solid, dashed, or animated.", None, ""),
 "MvDdsp": ("How the Move tool previews the object: ghost, outline, or off.", None, ""),
 "RsmCmp": (".rsm save format: uncompressed, LZ4, or zstd.", None, ""),
 "RsmBak": ("Keep a .rsm.bak backup copy each time you save.", None, ""),

 # --- Code-Audit Hardcoded (values currently baked into the source) ---
 "DefDClr": ("Default colour for newly drawn objects.", "colour", "hardcoded today"),
 "SelClr":  ("Highlight colour for selected objects.", "colour", "hardcoded today"),
 "SnpSrcClr":("Highlight colour for the object a snap is coming from.", "colour", "hardcoded today"),
 "SnpClr":  ("Colour of snap markers and their labels.", "colour", "hardcoded today"),
 "IntClr":  ("Colour of intersection markers.", "colour", "DUPLICATE of IntsCol"),
 "ExtClr":  ("Colour of the dashed imaginary-extension guide lines.", "colour", "hardcoded today"),
 "PreClr":  ("Colour of the preview / rubber-band line.", "colour", "hardcoded today"),
 "ExtSpd":  ("How fast the extension-guide dashes drift (animation).", "px/sec", "hardcoded today"),
 "ExtFade": ("Base transparency of the extension-guide dashes.", "0-1", "hardcoded today"),
 "ExtDshL": ("Dash length of the extension guide lines.", "px", "hardcoded today"),
 "ExtGapL": ("Gap length between extension-guide dashes.", "px", "hardcoded today"),
 "WinDshSpd":("How fast the selection-window dashes drift.", "px/sec", "hardcoded today"),
 "SelDshClr":("Colour of the dashed selection-window overlay.", "colour", "hardcoded today"),
 "SelDshW": ("Line thickness of the selection-window dashes.", "px", "hardcoded today"),
 "SelDshL": ("Dash length of the selection-window overlay.", "px", "hardcoded today"),
 "SelDshG": ("Gap length between selection-window dashes.", "px", "hardcoded today"),
 "SelPlsMin":("Minimum brightness of the selection box's pulsing fill.", "0-1", "hardcoded today"),
 "SelPlsMax":("Maximum brightness of the selection box's pulsing fill.", "0-1", "hardcoded today"),
 "SelPlsHz": ("How fast the selection box pulses (cycles per second).", "Hz", "hardcoded today"),
 "HitTolPx": ("How close (pixels) a click must be to select an object.", "px", "hardcoded today"),
 "IntRad":  ("Search radius (pixels) when clicking to find an intersection.", "px", "hardcoded today"),
 "PairLim": ("Safety cap on how many object pairs are tested for intersections.", "count", "hardcoded today"),
 "TabCycR": ("How far the cursor must move before Tab-cycle restarts.", "px", "hardcoded today"),
 "ArrCol":  ("Default number of columns in a rectangular array.", "count", "hardcoded today"),
 "ArrRow":  ("Default number of rows in a rectangular array.", "count", "hardcoded today"),
 "ArrDX":   ("Default column spacing in an array.", "units", "hardcoded today"),
 "ArrDY":   ("Default row spacing in an array.", "units", "hardcoded today"),
 "DfltZm":  ("Default zoom scale on startup.", "factor", "hardcoded today"),
 "DemoOn":  ("Load demo objects when the app starts.", None, "dev/demo flag"),
 "GpuRgWd": ("Ring thickness when drawing circles on the GPU.", "px", "hardcoded today"),
 "TessCirc":("How finely circles are broken into segments (CPU).", "factor", "hardcoded today"),
 "TessArc": ("How finely arcs are broken into segments.", "factor", "hardcoded today"),
 "TessEll": ("How finely ellipses are broken into segments.", "factor", "hardcoded today"),
 "TessEArc":("How finely elliptical arcs are broken into segments.", "factor", "hardcoded today"),

 # --- Grid & CARD ---
 "GrdEnb": ("Show the background reference grid.", None, ""),
 "GrdSnp": ("Snap the cursor onto grid intersections as you draw.", None, ""),
 "GrdSpc": ("Spacing between grid lines, in drawing units.", "units", ""),
 "CrdEnb": ("CARD lock: constrain drawing to cardinal directions (horizontal / vertical).", None, ""),

 # --- UCS Icon ---
 "UcsIcn": ("Show the coordinate-system (UCS) indicator icon.", None, ""),
 "UcsMod": ("Where the UCS icon sits: a screen corner, or at the origin.", None, ""),
 "UcsAvP": ("Image file used for the UCS X-axis avatar.", "path", ""),

 # --- Drafting Defaults ---
 "FltRad": ("Default radius used by the Fillet command.", "units", ""),
 "ChmDs1": ("Default first distance used by the Chamfer command.", "units", ""),
 "ChmDs2": ("Default second distance used by the Chamfer command.", "units", ""),
 "OfsDis": ("Default distance used by the Offset command.", "units", ""),
 "WlThk":  ("Default thickness for walls.", "units", ""),
 "TxHt":   ("Default height for new text.", "units", ""),
 "WlCnL":  ("Show the centerline of walls.", None, ""),
 "TrmMd":  ("Trim mode shared by Fillet and Chamfer: trim the originals, or keep them.", None, ""),
}

# ---- colours + sketches ---------------------------------------------------
BG=(27,33,43,255); GREY=(200,210,222,255); CROSS=(174,185,200,255); SEL=(255,200,80,255)
HOVER=(120,240,255,255); BLUE=(120,170,255,255); GREEN=(120,230,120,255); RED=(255,90,90,255)
GHOST=(255,255,255,95); FADE=(90,100,114,255); MASK=(232,237,243,235); MAG=(210,140,255,255); ARROW=(154,163,178,255)
W,H=160,100
def dline(d,p0,p1,fill,width=1,dash=5,gap=3):
    x0,y0=p0;x1,y1=p1;dx,dy=x1-x0,y1-y0;L=math.hypot(dx,dy)
    if L==0:return
    ux,uy=dx/L,dy/L;pos=0.0
    while pos<L:
        e=min(pos+dash,L);d.line([(x0+ux*pos,y0+uy*pos),(x0+ux*e,y0+uy*e)],fill=fill,width=width);pos+=dash+gap
def drect_dashed(d,box,fill,width=1):
    x0,y0,x1,y1=box
    dline(d,(x0,y0),(x1,y0),fill,width);dline(d,(x1,y0),(x1,y1),fill,width)
    dline(d,(x1,y1),(x0,y1),fill,width);dline(d,(x0,y1),(x0,y0),fill,width)
def cv():
    im=Image.new("RGBA",(W,H),BG);return im,ImageDraw.Draw(im)
def s_crosshair():
    im,d=cv();d.line([(15,50),(45,50)],CROSS);d.line([(30,35),(30,65)],CROSS)
    d.line([(70,50),(152,50)],CROSS);d.line([(105,16),(105,84)],CROSS);return im
def s_aperture():
    im,d=cv();d.line([(28,52),(122,52)],CROSS);d.line([(75,20),(75,84)],CROSS)
    d.line([(18,82),(132,26)],GREY);d.rectangle([58,36,92,68],outline=HOVER);d.rectangle([71,48,79,56],outline=MAG);return im
def s_window():
    im,d=cv();d.line([(40,40),(70,70)],GREY);d.line([(48,72),(95,45)],GREY)
    d.rectangle([35,32,118,80],fill=(120,170,255,45),outline=BLUE);return im
def s_crossing():
    im,d=cv();d.line([(40,40),(70,70)],GREY);d.line([(48,72),(95,45)],GREY)
    d.rectangle([35,32,118,80],fill=(120,230,120,40));drect_dashed(d,(35,32,118,80),GREEN);return im
def s_highlight():
    im,d=cv();d.line([(20,36),(80,36)],GREY);d.line([(82,70),(150,70)],SEL,width=3)
    d.rectangle([79,67,85,73],fill=SEL);d.rectangle([147,67,153,73],fill=SEL);return im
def s_hover():
    im,d=cv();d.line([(18,62),(120,36)],HOVER,width=3);d.line([(18,82),(96,82)],GREY)
    d.polygon([(86,44),(104,52),(96,55),(102,66),(98,68),(91,57),(86,58)],fill=(230,235,242,255));return im
def s_intersection():
    im,d=cv();d.line([(18,30),(92,84)],GREY);d.line([(18,84),(92,30)],GREY);d.line([(84,80),(150,38)],GREY)
    d.ellipse([51,52,61,62],outline=RED,width=2);d.ellipse([100,53,110,63],outline=RED,width=2);return im
def s_ghost():
    im,d=cv();d.rectangle([22,34,72,70],outline=GREY);drect_dashed(d,(88,52,138,88),GHOST);dline(d,(74,52),(96,70),ARROW);return im
def s_fade():
    im,d=cv();d.rectangle([20,32,70,80],outline=GREY,width=2);d.rectangle([95,32,145,80],outline=FADE,width=1);return im
def s_wipeout():
    im,d=cv();d.line([(18,62),(100,34)],GREY);d.line([(18,40),(100,72)],GREY);d.rectangle([34,38,96,72],fill=MASK,outline=SEL);return im
def s_transparency():
    im,d=cv();d.rectangle([20,30,72,82],fill=(120,170,255,255));d.rectangle([90,30,144,82],fill=(120,170,255,95));d.line([(82,72),(150,46)],RED);return im
def s_cloud():
    im,d=cv()
    for cx,cy in [(54,46),(74,40),(98,40),(116,52),(110,72),(86,80),(60,76),(44,60)]:
        d.ellipse([cx-15,cy-13,cx+15,cy+13],outline=SEL,width=2)
    return im
GRIPC=(64,153,255,255); GRIPH=(255,100,100,255)   # cool / hot grip
def _grip(d,x,y,h=5,c=GRIPC):
    d.rectangle([x-h,y-h,x+h,y+h],fill=c,outline=(20,30,45,255))
def s_grips():            # cool grips on a line (GrpEnb / GrClrU)
    im,d=cv();d.line([(22,55),(138,55)],GREY)
    for x in (24,80,136): _grip(d,x,55)
    return im
def s_grip_hot():         # one hot grip among cool (GrClrS)
    im,d=cv();d.line([(22,55),(138,55)],GREY)
    _grip(d,24,55);_grip(d,80,55,6,GRIPH);_grip(d,136,55);return im
def s_grip_size():        # small vs large grip (GrpSz)
    im,d=cv();d.line([(20,72),(140,72)],GREY)
    _grip(d,44,46,4);_grip(d,104,46,9)
    d.line([(40,86),(48,86)],GREY);d.line([(94,86),(114,86)],GREY);return im
def s_grip_radius():      # grab zone around a grip (GrpHvR)
    im,d=cv();d.line([(22,55),(138,55)],GREY)
    d.ellipse([60,35,100,75],outline=HOVER);_grip(d,80,55);return im
def s_cycle():            # overlapping objects, cycle badge (SelCyc)
    im,d=cv();d.rectangle([34,34,92,74],outline=GREY);d.rectangle([62,50,120,86],outline=HOVER)
    d.ellipse([100,26,120,46],outline=SEL,width=2);d.line([(110,30),(116,36)],SEL);return im
def s_pickbox():          # pick box at the crosshair over a line (PkBxSz)
    im,d=cv();d.line([(20,74),(140,40)],GREY)
    d.line([(58,57),(102,57)],CROSS);d.line([(80,35),(80,79)],CROSS)
    d.rectangle([70,47,90,67],outline=HOVER);return im
def s_polar():            # polar tracking rays, one locked (PolAng/PolMod/PolAdA)
    im,d=cv();cx,cy=38,76
    for a in (0,45,90,135):
        x=cx+72*math.cos(math.radians(a)); y=cy-72*math.sin(math.radians(a)); dline(d,(cx,cy),(x,y),GREY)
    x=cx+72*math.cos(math.radians(45)); y=cy-72*math.sin(math.radians(45)); d.line([(cx,cy),(x,y)],HOVER,width=2)
    d.ellipse([cx-3,cy-3,cx+3,cy+3],fill=SEL);return im
def s_polardist():        # distance ticks along a ray, cursor on a tick (PolDst)
    im,d=cv();cx,cy=24,72;x1,y1=140,30
    dx,dy=x1-cx,y1-cy;L=math.hypot(dx,dy);ux,uy=dx/L,dy/L;nx,ny=-uy,ux
    d.line([(cx,cy),(x1,y1)],GREY)
    for t in range(1,6):
        px=cx+ux*(L*t/6);py=cy+uy*(L*t/6);d.line([(px-5*nx,py-5*ny),(px+5*nx,py+5*ny)],CROSS)
    px=cx+ux*(L*4/6);py=cy+uy*(L*4/6);d.ellipse([px-4,py-4,px+4,py+4],fill=SEL);return im
def s_edgemode():         # trim/extend to an imaginary extension (EdgMod)
    im,d=cv();d.line([(95,18),(95,86)],GREY)            # cutting edge
    d.line([(20,55),(70,55)],GREY,width=2)              # object stops short
    dline(d,(70,55),(95,55),HOVER)                      # dashed extension to edge
    d.ellipse([90,50,100,60],outline=SEL,width=2);return im
def s_hatchsep():         # two separately-hatched areas (HpSep)
    im,d=cv()
    for x0 in (24,86):
        x1,y0,y1=x0+44,34,76;d.rectangle([x0,y0,x1,y1],outline=GREY)
        for o in range(8,44,9): d.line([(x0,y0+o),(x0+o,y0)],GREY); d.line([(x1-o,y1),(x1,y1-o)],GREY)
    return im
def s_autosave():         # clock + disk (SavTim)
    im,d=cv();d.ellipse([34,28,86,80],outline=GREY,width=2)
    d.line([(60,54),(60,38)],HOVER,width=2);d.line([(60,54),(74,60)],HOVER,width=2)
    d.rectangle([100,44,130,76],outline=SEL);d.rectangle([106,44,124,55],fill=SEL);return im
def s_backup():           # current doc over a faded backup copy (IsvBak)
    im,d=cv();d.rectangle([34,36,84,86],outline=FADE)
    d.rectangle([56,24,106,74],fill=BG,outline=GREY)
    for y in (34,44,54,64): d.line([(64,y),(98,y)],GREY)
    return im
def s_raster():           # framed raster image (RastDpi)
    im,d=cv();d.rectangle([30,26,130,82],outline=GREY)
    d.ellipse([44,36,60,52],outline=SEL)
    d.line([(32,80),(64,52),(86,70),(106,48),(128,80)],fill=GREY);return im
def s_menubar():          # window with a classic menu strip (MnuBar)
    im,d=cv();d.rectangle([20,22,140,84],outline=GREY);d.rectangle([20,22,140,37],outline=HOVER)
    for x in (28,52,76,100): d.line([(x,29),(x+14,29)],HOVER,width=2)
    return im
def s_tooltip():          # button with a tooltip popup (TltEnb)
    im,d=cv();d.rectangle([38,28,64,54],outline=HOVER);d.line([(46,41),(56,41)],HOVER,width=2)
    d.rectangle([58,60,128,78],fill=(40,48,62,255),outline=GREY);d.line([(64,69),(122,69)],GREY);return im
def s_textfill():         # filled vs outline glyph (TxtFil/TxtQlt)
    im,d=cv();d.polygon([(40,74),(53,32),(66,74)],fill=GREY);d.line([(45,60),(61,60)],BG,width=2)
    d.line([(86,74),(99,32),(112,74)],GREY,width=2);d.line([(91,60),(107,60)],GREY,width=2);return im
def s_smoothness():       # coarse polygon vs smooth circle (WhipArc)
    im,d=cv();cx,cy,r=46,55,22
    pts=[(cx+r*math.cos(math.radians(a)),cy-r*math.sin(math.radians(a))) for a in range(0,360,60)]
    d.polygon(pts,outline=GREY);d.ellipse([90,33,134,77],outline=HOVER);return im
def s_zoom():             # magnifier + wheel zoom (ZmFact/ZmWhl)
    im,d=cv();d.ellipse([42,28,82,68],outline=HOVER,width=2);d.line([(78,64),(98,84)],HOVER,width=3)
    d.line([(54,48),(70,48)],HOVER,width=2);d.line([(62,40),(62,56)],HOVER,width=2)
    d.line([(112,34),(112,74)],GREY);d.line([(108,40),(112,34),(116,40)],SEL);d.line([(108,68),(112,74),(116,68)],SEL);return im
def s_rubberband():       # dashed band from anchor to cursor (RubBnd)
    im,d=cv();d.ellipse([28,68,38,78],fill=SEL);dline(d,(33,73),(108,36),HOVER)
    d.line([(94,36),(122,36)],CROSS);d.line([(108,22),(108,50)],CROSS);return im
def s_canvas():           # canvas background swatch (BgCol)
    im,d=cv();d.rectangle([34,26,126,82],fill=(18,22,28,255),outline=GREY)
    d.line([(48,68),(80,40),(112,66)],fill=GREY);return im
def s_array():            # grid of copies (ArrCol/ArrRow/ArrDX/ArrDY)
    im,d=cv()
    for rr in range(3):
        for cc in range(4):
            x=28+cc*26;y=30+rr*18;d.rectangle([x,y,x+16,y+12],outline=GREY)
    d.rectangle([28,30,44,42],outline=SEL);return im
def s_grid():             # reference grid (GrdEnb/GrdSpc)
    im,d=cv()
    for x in range(28,134,18): d.line([(x,22),(x,86)],FADE)
    for y in range(22,88,18): d.line([(28,y),(132,y)],FADE)
    return im
def s_gridsnap():         # cursor locked to a grid node (GrdSnp)
    im,d=cv()
    for x in range(28,134,18): d.line([(x,22),(x,86)],FADE)
    for y in range(22,88,18): d.line([(28,y),(132,y)],FADE)
    px,py=82,58;d.line([(px-12,py),(px+12,py)],HOVER);d.line([(px,py-12),(px,py+12)],HOVER)
    d.ellipse([px-4,py-4,px+4,py+4],fill=SEL);return im
def s_cardinal():         # H/V cardinal lock (CrdEnb)
    im,d=cv();cx,cy=80,55
    d.line([(cx,20),(cx,90)],HOVER,width=2);d.line([(40,cy),(120,cy)],HOVER,width=2)
    d.line([(cx-4,26),(cx,20),(cx+4,26)],HOVER);d.line([(cx-4,84),(cx,90),(cx+4,84)],HOVER)
    d.line([(46,cy-4),(40,cy),(46,cy+4)],HOVER);d.line([(114,cy-4),(120,cy),(114,cy+4)],HOVER);return im
def s_fillet():           # rounded corner (FltRad)
    im,d=cv();d.line([(34,28),(34,64)],GREY,width=2)
    d.arc([34,44,74,84],90,180,fill=SEL,width=3);d.line([(54,84),(124,84)],GREY,width=2);return im
def s_chamfer():          # bevelled corner (ChmDs1/ChmDs2)
    im,d=cv();d.line([(34,28),(34,62)],GREY,width=2)
    d.line([(34,62),(62,84)],SEL,width=3);d.line([(62,84),(124,84)],GREY,width=2);return im
def s_offset():           # parallel copy at a distance (OfsDis)
    im,d=cv();d.line([(28,42),(124,42)],GREY,width=2);d.line([(28,62),(124,62)],HOVER,width=2)
    d.line([(40,42),(40,62)],ARROW);return im
def s_wall():             # double line + centerline (WlThk/WlCnL)
    im,d=cv();d.line([(28,42),(130,42)],GREY,width=2);d.line([(28,68),(130,68)],GREY,width=2)
    dline(d,(28,55),(130,55),FADE);return im
def s_ucs():              # X/Y axis icon (UcsIcn/UcsMod)
    im,d=cv();ox,oy=46,74
    d.line([(ox,oy),(ox,32)],HOVER,width=3);d.line([(ox,oy),(108,oy)],SEL,width=3)
    d.line([(ox-4,40),(ox,32),(ox+4,40)],HOVER);d.line([(100,oy-4),(108,oy),(100,oy+4)],SEL);return im
SKETCH={"CrsHrS":s_crosshair,"AperBx":s_aperture,"WinACol":s_window,"CrsACol":s_crossing,
        "HltSel":s_highlight,"SelPrv":s_hover,"IntsCol":s_intersection,"IntsDsp":s_intersection,
        "DrDspM":s_ghost,"LyLkFd":s_fade,"WpFrmM":s_wipeout,"TrnDsp":s_transparency,"RvClCrM":s_cloud,
        # Selection & Grips
        "GrpEnb":s_grips,"GrClrU":s_grips,"GrClrS":s_grip_hot,"GrpSz":s_grip_size,
        "GrpHvR":s_grip_radius,"SelCyc":s_cycle,"PkAuto":s_window,"PkDrag":s_window,
        # Object Snaps & Precision
        "PkBxSz":s_pickbox,"SpTGSZ":s_aperture,"PolAng":s_polar,"PolMod":s_polar,
        "PolAdA":s_polar,"PolDst":s_polardist,
        # Editing & Behavior
        "EdgMod":s_edgemode,"HpSep":s_hatchsep,"XFdCtl":s_fade,
        # File & Save
        "SavTim":s_autosave,"IsvBak":s_backup,
        # Xrefs & Images
        "XdwFd":s_fade,"RastDpi":s_raster,
        # UI & Workspace
        "MnuBar":s_menubar,"TltEnb":s_tooltip,"PalOpq":s_transparency,
        # System & Performance
        "TxtFil":s_textfill,"TxtQlt":s_textfill,"WhipArc":s_smoothness,
        # View & Navigation
        "ZmFact":s_zoom,"ZmWhl":s_zoom,
        # RUST_CAD-specific
        "RubBnd":s_rubberband,"MvDdsp":s_ghost,"BgCol":s_canvas,
        # Code-Audit Hardcoded
        "IntClr":s_intersection,"PreClr":s_rubberband,"SelClr":s_highlight,"ExtClr":s_edgemode,
        "SelDshClr":s_window,"ArrCol":s_array,"ArrRow":s_array,"ArrDX":s_array,"ArrDY":s_array,
        "TessCirc":s_smoothness,"TessArc":s_smoothness,"TessEll":s_smoothness,"TessEArc":s_smoothness,
        # Grid & CARD / UCS Icon / Drafting Defaults
        "GrdEnb":s_grid,"GrdSpc":s_grid,"GrdSnp":s_gridsnap,"CrdEnb":s_cardinal,
        "UcsIcn":s_ucs,"UcsMod":s_ucs,
        "FltRad":s_fillet,"ChmDs1":s_chamfer,"ChmDs2":s_chamfer,"OfsDis":s_offset,
        "WlThk":s_wall,"WlCnL":s_wall}

# ---- known duplicates / overlaps (decide later; just flag for now) ---------
# kind: "dup" = same thing / explicit alias (red);  "overlap" = related, partial (amber)
OVERLAP = {
 # --- true duplicates / aliases ---
 "IntsCol": ("dup",     "IntClr (Code-Audit Hardcoded) is an explicit alias - same intersection colour."),
 "IntClr":  ("dup",     "Alias of IntsCol (Display). Duplicate."),
 # --- partial overlaps across sections ---
 "HltSel":  ("overlap", "SelClr (Code-Audit) is the selected-object highlight colour - overlaps this toggle's effect."),
 "SelClr":  ("overlap", "Overlaps HltSel (Display) - the highlight colour for the same effect."),
 "AperBx":  ("overlap", "Pairs with SpTGSZ (Object Snaps), which sets the box SIZE. Snap-box family - consider co-locating."),
 "DrDspM":  ("overlap", "MvDdsp & RubBnd (RUST-specific) also control the drag / move-ghost rendering."),
 "WinACol": ("overlap", "SelDshClr / SelDshW... (Code-Audit) style the WINDOW box's dashed overlay - same box."),
 "CrsACol": ("overlap", "Crossing box; SelDsh* / SelPls* (Code-Audit) animate the same selection basket."),
 "SelAr":   ("overlap", "SelPls* / SelDsh* (Code-Audit) actually render this selection-area pulse/dash effect."),
 "IntsDsp": ("overlap", "IntRad (Code-Audit) = the click search radius for these same intersection markers."),
 "TrkPth":  ("overlap", "ExtClr / ExtSpd / ExtDshL... (Code-Audit) style the tracking / extension dashed guides."),
 "EdgMod":  ("overlap", "Related to TrmMd (Drafting Defaults) - both govern Trim/Extend behaviour."),
 "XdwFd":   ("overlap", "Fade-% family with LyLkFd (Display) and XFdCtl (Editing) - 3 separate fade settings."),
 "TltEnb":  ("overlap", "Tooltip family: RllTp (object hover) & GrpTip (grip) are separate tooltip toggles."),
 "WhipArc": ("overlap", "TessCirc / TessArc / TessEll (Code-Audit) are the actual per-curve smoothness factors."),
 "FlDlgM":  ("overlap", "Dialog-suppression family with CmDlgM (Editing) - both swap dialogs for prompts."),
 # --- snap-family overlaps (appear once those sections are added) ---
 "PkBxSz":  ("overlap", "HitTolPx (Code-Audit) - desc literally says 'overlaps PkBxSz'."),
 "SpTGSZ":  ("overlap", "AperBx (Display) toggles this same snap-target box; this sets its size."),
 "HitTolPx":("overlap", "Overlaps PkBxSz (Object Snaps) - hit-test tolerance, same idea."),
 "OsnOpt":  ("overlap", "Snap var stranded in Selection & Grips - belongs with Object Snaps."),
 "OsnNdLg": ("overlap", "Snap var stranded in Selection & Grips - belongs with Object Snaps."),
 "TabCyc":  ("overlap", "TabCycR (Code-Audit) = px before this Tab-cycle resets. Same feature."),
 "TabCycR": ("overlap", "Tied to TabCyc (RUST-specific) - the cursor-move reset distance."),
 "MvDdsp":  ("overlap", "Overlaps DrDspM (Display) - both the move/drag ghost preview. Pick one model."),
 "RubBnd":  ("overlap", "PreClr (Code-Audit) sets this rubber-band's colour - style here, colour there."),
 "SnpPri":  ("overlap", "Snap config - sits apart from the Object Snaps section (with SnpAct/OsnOpt family)."),
 "SnpAct":  ("overlap", "Snap config - belongs with the Object Snaps section (OsnOpt/SnpPri family)."),
 "PreClr":  ("overlap", "Sets the colour of RubBnd (RUST-specific) - colour here, style there."),
 "SnpClr":  ("overlap", "Snap-marker colour - part of the snap-display family (with SnpSrcClr / Object Snaps)."),
 "SnpSrcClr":("overlap","Snap-source highlight colour - snap-display family (with SnpClr)."),
 "ExtClr":  ("overlap", "Styles the tracking / extension guides toggled by TrkPth (Display)."),
 "SelDshClr":("overlap","Renders the selection-window box that WinACol / CrsACol (Display) colour."),
 "TrmMd":   ("overlap", "Shared by Fillet & Chamfer; related to EdgMod (Editing) trim/extend behaviour."),
}

# ---- build workbook -------------------------------------------------------
wb=Workbook(); ws=wb.active; ws.title="Settings Guide"
HEADERS=["Section","Variable","Sketch","What it does","Type","Options / Range","Default","Status","Editable now?","Notes","Overlap / Duplicate"]
hfill=PatternFill("solid",fgColor="223046"); hfont=Font(bold=True,color="E6EBF2",size=11)
ws.append(HEADERS)
for c in range(1,len(HEADERS)+1):
    cell=ws.cell(1,c);cell.fill=hfill;cell.font=hfont
    cell.alignment=Alignment(vertical="center",horizontal="center",wrap_text=True)
for i,w in enumerate([22,11,24,52,12,24,12,10,11,22,40],1):
    ws.column_dimensions[get_column_letter(i)].width=w
SFILL={"Active":"1d3a2b","Planned":"1e3146","Stub":"2a2d34","Tentative":"3a3320"}
SFONT={"Active":"43B581","Planned":"5B9BD5","Stub":"9AA3B2","Tentative":"D4A853"}
# duplicate/overlap cell styling
DUPFILL=PatternFill("solid",fgColor="5A2330");  DUPFONT=Font(name="Consolas",bold=True,color="FF9090")
OVLFILL=PatternFill("solid",fgColor="3A3320");  OVLFONT=Font(name="Consolas",bold=True,color="E0C060")

allvars=parse_varreg()
r=2; nsk=0
for v in allvars:
    if v["sec"] not in SECTIONS: continue
    typ,opt_d,opts=kind_type_options(v["kraw"])
    ex=EXPLAIN.get(v["name"])
    what = ex[0] if ex else v["desc"]
    opt  = (ex[1] if (ex and ex[1]) else opt_d)
    notes= (ex[2] if ex else "")
    dflt = default_display(v["default"], opts)
    ov = OVERLAP.get(v["name"])
    ws.cell(r,1,v["sec"]);
    nc=ws.cell(r,2,v["name"]); nc.font=Font(name="Consolas",bold=True)
    ws.cell(r,4,what).alignment=Alignment(wrap_text=True,vertical="center")
    ws.cell(r,5,typ); ws.cell(r,6,opt); ws.cell(r,7,dflt)
    sc=ws.cell(r,8,v["status"]); sc.fill=PatternFill("solid",fgColor=SFILL.get(v["status"],"2a2d34"))
    sc.font=Font(bold=True,color=SFONT.get(v["status"],"9AA3B2")); sc.alignment=Alignment(horizontal="center")
    ws.cell(r,9,"Yes" if v["wired"] else "No").alignment=Alignment(horizontal="center")
    ws.cell(r,10,notes)
    oc=ws.cell(r,11,ov[1] if ov else "")
    oc.alignment=Alignment(wrap_text=True,vertical="center")
    if ov:
        kind=ov[0]
        nc.fill = DUPFILL if kind=="dup" else OVLFILL
        nc.font = DUPFONT if kind=="dup" else OVLFONT
        oc.font = Font(color="FF9090" if kind=="dup" else "E0C060", italic=True)
    for c in (1,5,6,7): ws.cell(r,c).alignment=Alignment(vertical="center")
    if v["name"] in SKETCH:
        p=os.path.join(PNG,f'{v["name"]}.png'); SKETCH[v["name"]]().save(p)
        xi=XLImage(p); xi.width=150; xi.height=94; ws.add_image(xi,f"C{r}")
        ws.row_dimensions[r].height=74; nsk+=1
    else:
        ws.row_dimensions[r].height=30
    r+=1
ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:K{r-1}"

# ---- Legend sheet ---------------------------------------------------------
lg=wb.create_sheet("Legend")
lg.column_dimensions["A"].width=14; lg.column_dimensions["B"].width=92
lg.append(["Status","What it means"])
lg.cell(1,1).font=Font(bold=True,color="E6EBF2"); lg.cell(1,1).fill=hfill
lg.cell(1,2).font=Font(bold=True,color="E6EBF2"); lg.cell(1,2).fill=hfill
LEG=[("ACTIVE","Built and working - change it and the app behaves differently right away.","Active"),
     ("PLANNED","We intend to build it. The setting is reserved; some are already editable & saved (your value is stored, just not used by a feature yet), others are shown but disabled.","Planned"),
     ("STUB","Copied from AutoCAD's master list for compatibility, but NOT planned here - usually a feature this 2D app doesn't have (3D, printing, OLE, point clouds, lights, tray...). Easiest to delete.","Stub"),
     ("TENTATIVE","Kept but we're unsure we need it - revisit later (promote or remove).","Tentative")]
for i,(s,m,key) in enumerate(LEG,2):
    a=lg.cell(i,1,s); a.font=Font(bold=True,color=SFONT[key]); a.fill=PatternFill("solid",fgColor=SFILL[key]); a.alignment=Alignment(horizontal="center")
    lg.cell(i,2,m).alignment=Alignment(wrap_text=True,vertical="center"); lg.row_dimensions[i].height=44
lg.cell(7,1,"Editable now?"); lg.cell(7,1).font=Font(bold=True)
lg.cell(7,2,"Yes = you can change & save it today (it maps to a real saved setting). No = catalogued for later; shown but disabled in the panel.")
lg.cell(9,1,"Sketch colours"); lg.cell(9,1).font=Font(bold=True)
lg.cell(9,2,"yellow = selected/highlight  |  cyan = hover preview / aperture  |  blue = window box (L->R)  |  green = crossing box (R->L)  |  red = intersection marker")
# overlap / duplicate legend
od=lg.cell(11,1,"Overlap / Duplicate"); od.font=Font(bold=True)
lg.cell(11,2,"Variable name is highlighted when it clashes with another variable. Decision deferred - just flagged for now.")
dl=lg.cell(12,1,"Duplicate"); dl.fill=DUPFILL; dl.font=DUPFONT; dl.alignment=Alignment(horizontal="center")
lg.cell(12,2,"RED = same thing / explicit alias (e.g. IntClr = IntsCol). One should be deleted once we pick the survivor.")
ol=lg.cell(13,1,"Overlap"); ol.fill=OVLFILL; ol.font=OVLFONT; ol.alignment=Alignment(horizontal="center")
lg.cell(13,2,"AMBER = related / partial overlap (same feature from a different angle, or a var stranded in the wrong section). Revisit when we split sections.")
for i in (11,12,13): lg.row_dimensions[i].height=30

out=os.path.join(ROOT,"AutoRASM_Settings_Guide.xlsx"); wb.save(out)
print("wrote",out,"| rows:",r-2,"| sketches:",nsk,"| order = varreg.rs")
